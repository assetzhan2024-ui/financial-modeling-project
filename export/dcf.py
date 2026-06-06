"""
export/dcf.py
=============
DCF Valuation Excel — по шаблону NBA_ADVANCED DCF model v2.

Three sheets per ticker:
  "Company fin forecasts"
  "DCF input"
  "DCF output"

Key improvements over v1:
  - Per-year revenue growth inputs (editable, default tapering from base)
  - EBITDA margin converges from actual base to LT target over 10 years
  - All year-level assumptions editable directly in the sheet
  - Cross-sheet formula names carry suffix → no #REF for multi-ticker
  - _safe_ratio guards against bad Yahoo data (negative D&A, zero EBITDA)

Публичный API:
    build_dcf_multi(records, fund_map=None) -> bytes
"""

import io
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError:
    raise ImportError("pip install openpyxl")


# ── Palette ────────────────────────────────────────────────────────────────
_BG_DARK   = "1F2D3D"
_BG_HEAD   = "2C3E50"
_BG_INPUT  = "D6EAF8"   # blue  — editable
_BG_CALC   = "FDFEFE"   # white — formula
_BG_OUT    = "FDEBD0"   # amber — output
_BG_SECT   = "EBF5FB"   # light blue — sub-headers
_BG_ALT    = "F4F6F7"
_FG_W      = "FFFFFF"
_FG_D      = "1C2833"
_FG_G      = "717D7E"
_FG_BLUE   = "1A5276"
_BRD       = "BFC9CA"

_FMT_NUM   = '#,##0.0'
_FMT_PCT   = '0.0%'
_FMT_PRICE = '#,##0.00'


def _bd():
    s = Side(style="thin", color=_BRD)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)
def _f(sz=9, bold=False, color=_FG_D):
    return Font(name="Calibri", size=sz, bold=bold, color=color)
def _w(col: int) -> str:
    return get_column_letter(col)

def _cell(ws, row, col, value, bg=_BG_CALC, bold=False,
          color=_FG_D, fmt=None, align="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", size=9, bold=bold, color=color)
    c.fill      = _fill(bg)
    c.border    = _bd()
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        c.number_format = fmt
    return c

def _lbl(ws, row, col, text, bg=_BG_SECT, indent=0):
    c = ws.cell(row=row, column=col, value=(" " * indent) + str(text))
    c.font      = _f(9, False, _FG_G)
    c.fill      = _fill(bg)
    c.border    = _bd()
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def _inp(ws, row, col, value, fmt=None):
    return _cell(ws, row, col, value, bg=_BG_INPUT, color=_FG_BLUE, fmt=fmt)

def _out(ws, row, col, value, fmt=None):
    return _cell(ws, row, col, value, bg=_BG_OUT, bold=True, fmt=fmt)

def _hdr(ws, row, col, text, end_col=None, bg=_BG_HEAD, sz=10):
    if end_col and end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", size=sz, bold=True, color=_FG_W)
    c.fill      = _fill(bg)
    c.border    = _bd()
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c


# ── Data helpers ──────────────────────────────────────────────────────────

def _get_fund(fund: Optional[dict], section: str, yf_key: str,
              default=None) -> Optional[float]:
    """Latest-year value from fund_data, converted to millions."""
    if not fund:
        return default
    for yr in sorted(fund.get(section, {}).keys(), reverse=True)[:1]:
        v = fund.get(section, {}).get(yr, {}).get(yf_key)
        if v is not None:
            try:
                return float(v) / 1e6
            except (TypeError, ValueError):
                pass
    return default


def _safe_ratio(numerator, denominator, default=0.0,
                min_val=0.0, max_val=1.0) -> float:
    """
    Safely compute numerator/denominator, clamped to [min_val, max_val].
    Returns default when data is missing, zero, or out of range.
    Protects against bad yfinance data (negative D&A, zero EBITDA for
    financials/insurance, etc.).
    """
    try:
        if numerator is None or denominator is None or denominator == 0:
            return default
        ratio = float(numerator) / float(denominator)
        return max(min_val, min(max_val, ratio))
    except (TypeError, ValueError, ZeroDivisionError):
        return default


def _taper_growth(base: float, n: int = 10) -> list[float]:
    """
    Return a list of n annual growth rates that taper from `base`
    toward a long-run rate of 5% over the projection period.
    Simulates analyst practice of fading near-term growth.
    """
    lt = 0.05
    return [round(base + (lt - base) * i / (n - 1), 4) for i in range(n)]


def _converge_margin(base: float, n: int = 10) -> list[float]:
    """
    Return n margin values converging from `base` toward the long-run
    steady-state.  The long-run target is the midpoint between base and
    the 30% cap (or base itself if base > 30%, held flat).
    This produces a realistic gradual improvement / normalisation.
    """
    lt = min(0.35, base * 1.20) if base > 0 else 0.20
    return [round(base + (lt - base) * i / (n - 1), 6) for i in range(n)]


# ══════════════════════════════════════════════════════════════════════════
# Sheet 1: Company fin forecasts
# ══════════════════════════════════════════════════════════════════════════

def _build_forecasts(wb, rec: dict, fund: Optional[dict], suffix: str = ""):
    ws = wb.create_sheet(f"Company fin forecasts{suffix}")
    for col, w in [("A",28),("B",13),("C",13),("D",13),("E",13),("F",13)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22

    name = rec.get("name") or rec.get("ticker","X")
    cur  = rec.get("currency") or "USD"

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
                value=f"{name} — 5 Year Financial Forecast  [{cur}m]")
    c.font = Font(name="Calibri", size=12, bold=True, color=_FG_W)
    c.fill = _fill(_BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    inc   = (fund or {}).get("income",   {})
    years = sorted(inc.keys())[-5:]          # up to 5 years, oldest→newest
    fcols = list(range(2, 2 + len(years)))   # B..F

    for i, yr in enumerate(years):
        c2 = ws.cell(row=3, column=fcols[i], value=f"FY{yr[-2:]}")
        c2.font = Font(name="Calibri", size=10, bold=True, color=_FG_W)
        c2.fill = _fill(_BG_HEAD)
        c2.border = _bd()
        c2.alignment = Alignment(horizontal="center", vertical="center")

    def _row(rn, lbl, sect, key, pct_of=""):
        _lbl(ws, rn, 1, lbl)
        for i, yr in enumerate(years):
            raw  = (fund or {}).get(sect, {}).get(yr, {}).get(key)
            if raw is None:
                _cell(ws, rn, fcols[i], "—"); continue
            val = float(raw) / 1e6
            if pct_of:
                base = (fund or {}).get(sect, {}).get(yr, {}).get(pct_of)
                val  = float(raw) / float(base) if base else None
                _cell(ws, rn, fcols[i],
                      val if val is not None else "—", fmt=_FMT_PCT)
            else:
                _cell(ws, rn, fcols[i], val, fmt=_FMT_NUM)

    _hdr(ws, 4, 1, "Income Statement", end_col=6)
    _row(5,  "Revenue",            "income", "Total Revenue")

    # YoY growth
    _lbl(ws, 6, 1, "  YoY growth %")
    rvs = [(fund or {}).get("income",{}).get(yr,{}).get("Total Revenue")
           for yr in years]
    for i in range(len(years)):
        if i == 0 or not rvs[i] or not rvs[i-1]:
            _cell(ws, 6, fcols[i], "—")
        else:
            _cell(ws, 6, fcols[i], float(rvs[i])/float(rvs[i-1])-1, fmt=_FMT_PCT)

    _row(7,  "Gross Profit",       "income",   "Gross Profit")
    _row(8,  "  GM%",              "income",   "Gross Profit",
         pct_of="Total Revenue")
    _row(9,  "Operating Income",   "income",   "Operating Income")
    _row(10, "  Operating margin", "income",   "Operating Income",
         pct_of="Total Revenue")
    _row(11, "EBITDA",             "income",   "EBITDA")
    _row(12, "  EBITDA margin",    "income",   "EBITDA",
         pct_of="Total Revenue")
    _row(13, "Net Income",         "income",   "Net Income")
    _row(14, "  Net margin",       "income",   "Net Income",
         pct_of="Total Revenue")

    _hdr(ws, 16, 1, "Balance Sheet", end_col=6)
    _row(17, "Total Assets",       "balance",  "Total Assets")
    _row(18, "Total Debt",         "balance",  "Total Debt")
    _row(19, "Cash & Equivalents", "balance",  "Cash And Cash Equivalents")
    _row(20, "Equity",             "balance",  "Stockholders Equity")

    _hdr(ws, 22, 1, "Cash Flow Statement", end_col=6)
    _row(23, "Operating CF",       "cashflow", "Operating Cash Flow")
    _row(24, "Capex",              "cashflow", "Capital Expenditure")
    _row(25, "Free Cash Flow",     "cashflow", "Free Cash Flow")
    _row(26, "D&A",                "cashflow", "Depreciation And Amortization")

    ws.cell(row=28, column=1,
            value=f"Source: {name} annual reports  |  All values in {cur} millions"
            ).font = _f(8, False, _FG_G)


# ══════════════════════════════════════════════════════════════════════════
# Sheet 2: DCF input  — per-year growth + margin convergence
# ══════════════════════════════════════════════════════════════════════════
#
# Layout:
#   Rows 1-18  : title + key fixed assumptions (WACC, TGR, TV, tax, etc.)
#   Row 20     : section header "DCF FCF Calculation"
#   Row 21     : year labels   Base | Y+1 | Y+2 … Y+10
#   Row 22     : Revenue       (base = input, years = formula)
#   Row 23     : % growth      (BASE blank, Y+1..Y+10 = editable inputs)
#   Row 25     : EBITDA        (each year = Revenue × that year's margin)
#   Row 26     : EBITDA margin (each year = editable input, converges)
#   Row 27     : % growth      (computed)
#   Row 29     : D&A           (= −Revenue × $F$14)
#   Row 30     : EBIT          (= EBITDA + D&A)
#   Row 31     : EBIT margin
#   Row 33     : Tax           (= −EBIT × $F$17)
#   Row 35     : EBIAT         (= EBIT + Tax)
#   Row 37     : Add: D&A      (= −D&A, positive)
#   Row 39     : Capex         (= −Revenue × $F$15)
#   Row 40     : % of revenue
#   Row 42     : Change in NWC (= −Revenue × $F$16)
#   Row 44     : Unlevered FCF (= EBIAT + D&A + Capex + NWC)
#   Row 46     : Timing
#   Row 47     : Discount factor
#   Row 49-51  : Terminal Value

def _build_dcf_input(wb, rec: dict, fund: Optional[dict], suffix: str = ""):
    ws = wb.create_sheet(f"DCF input{suffix}")
    for col, w in [
        ("A",3),("B",3),("C",3),("D",26),("E",3),("F",12),
        ("G",3),("H",30),("I",12),("J",12),("K",12),("L",12),
        ("M",12),("N",12),("O",12),("P",12),("Q",12),("R",12),("S",12),
    ]:
        ws.column_dimensions[col].width = w

    name   = rec.get("name")   or rec.get("ticker","X")
    cur    = rec.get("currency") or "USD"
    region = rec.get("region",  "US")

    # ── Base data from fund ───────────────────────────────────────────────
    rev_m    = _get_fund(fund, "income",   "Total Revenue",       0.0) or 0.0
    ebitda_m = _get_fund(fund, "income",   "EBITDA",              None)
    dna_m    = _get_fund(fund, "cashflow", "Depreciation And Amortization", None)
    capex_m  = _get_fund(fund, "cashflow", "Capital Expenditure", None)
    debt_m   = _get_fund(fund, "balance",  "Total Debt",          0.0) or 0.0
    cash_m   = _get_fund(fund, "balance",  "Cash And Cash Equivalents", 0.0) or 0.0

    # Safe ratios — guard against zero/negative/missing Yahoo data
    ebitda_base = _safe_ratio(ebitda_m, rev_m, default=0.20,
                              min_val=0.0, max_val=0.90)
    dna_pct     = _safe_ratio(abs(dna_m)    if dna_m    else None, rev_m,
                              default=0.03, min_val=0.0, max_val=0.30)
    capex_pct   = _safe_ratio(abs(capex_m)  if capex_m  else None, rev_m,
                              default=0.05, min_val=0.0, max_val=0.30)

    wacc = {"KZ":0.14,"Emerging":0.12,"Europe":0.09,"Asia":0.10
            }.get(region, 0.094)
    tgr  = {"KZ":0.04,"Emerging":0.035}.get(region, 0.025)
    tv_m = {"KZ":6.0, "Emerging":7.0}.get(region, 8.0)

    # Per-year growth rates: taper from recent trend toward 5%
    recent_growth = 0.10   # default; will be overridden if fund data exists
    if fund:
        rvs = sorted(
            [(yr, float(v["Total Revenue"]))
             for yr, v in fund.get("income",{}).items()
             if v.get("Total Revenue")], key=lambda x: x[0])
        if len(rvs) >= 2:
            recent_growth = max(0.0, min(0.40,
                rvs[-1][1] / rvs[-2][1] - 1 if rvs[-2][1] else 0.10))

    growth_schedule = _taper_growth(recent_growth, 10)   # 10 values for Y+1..Y+10
    margin_schedule = _converge_margin(ebitda_base, 11)  # 11 values for base..Y+10

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells("D1:S1")
    c = ws.cell(row=1, column=4, value=f"DCF — {name}  [{cur}m]")
    c.font = Font(name="Calibri", size=12, bold=True, color=_FG_W)
    c.fill = _fill(_BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    _hdr(ws, 3, 4, "DCF", bg=_BG_HEAD)
    _hdr(ws, 5, 4, "Fixed assumptions (apply to all years unless overridden below)",
         bg=_BG_SECT, sz=9)

    def _assump(row, label, val, fmt=None, note=""):
        _lbl(ws, row, 4, label)
        _inp(ws, row, 6, val, fmt=fmt)
        if note:
            ws.cell(row=row, column=8, value=note).font = _f(8, False, _FG_G)

    _assump(7,  "Perpetuity growth rate (TGR)", tgr,    _FMT_PCT,
            "<<< Long-run GDP proxy — discuss with senior")
    _assump(8,  "TV exit EBITDA multiple",      tv_m,   _FMT_NUM,
            "<<< Cross-check with sector comps")
    _assump(9,  "WACC",                         wacc,   _FMT_PCT,
            "<<< CAPM-derived, region-adjusted")
    _assump(10, "D&A % of revenue",             dna_pct,    _FMT_PCT,
            f"<<< Latest year actual ({dna_pct:.1%}) — absolute value")
    _assump(11, "Capex % of revenue",           capex_pct,  _FMT_PCT,
            f"<<< Latest year actual ({capex_pct:.1%}) — absolute value")
    _assump(12, "Change in NWC % of revenue",   0.02,       _FMT_PCT)
    _assump(13, "Tax rate",                     0.20,       _FMT_PCT)

    ws.cell(row=15, column=4,
            value="↓  Per-year revenue growth and EBITDA margin are set individually below — edit directly in the blue cells"
            ).font = _f(8, True, _FG_BLUE)

    # ── Projection columns I..S (cols 9..19) ─────────────────────────────
    COLS = list(range(9, 20))   # I=9 (base), J..S = Y+1..Y+10
    N    = len(COLS)

    _hdr(ws, 18, 4, "DCF FCF Calculation — 10-year projection", end_col=19,
         bg=_BG_HEAD)

    # Year labels row 19
    for j, col in enumerate(COLS):
        label = "Base" if j == 0 else f"Y+{j}"
        c2 = ws.cell(row=19, column=col, value=label)
        c2.font = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        c2.fill = _fill(_BG_HEAD)
        c2.border = _bd()
        c2.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 21: Revenue ────────────────────────────────────────────────────
    _lbl(ws, 21, 4, "Revenue")
    _inp(ws, 21, 9, round(rev_m, 1), _FMT_NUM)     # I21 = base (editable)
    for j in range(1, N):
        col = COLS[j]
        # Revenue = prev_year_revenue × (1 + that_year_growth)
        _cell(ws, 21, col,
              f"={_w(col-1)}21*(1+{_w(col)}22)",
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 22: % growth (per-year editable inputs) ────────────────────────
    _lbl(ws, 22, 4, "  % growth  ← edit per year")
    # Base: blank (no growth rate for the starting point)
    ws.cell(row=22, column=9, value="—")
    for j in range(1, N):
        col = COLS[j]
        g   = growth_schedule[j - 1]     # j-1 because growth_schedule[0]=Y+1
        _inp(ws, 22, col, g, fmt=_FMT_PCT)

    # ── Row 24: EBITDA ─────────────────────────────────────────────────────
    _lbl(ws, 24, 4, "EBITDA")
    for j, col in enumerate(COLS):
        _cell(ws, 24, col,
              f"={_w(col)}21*{_w(col)}25",         # Revenue × that year's margin
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 25: EBITDA margin (per-year editable, converging) ──────────────
    _lbl(ws, 25, 4, "  EBITDA margin  ← edit per year")
    for j, col in enumerate(COLS):
        mg = margin_schedule[j]
        _inp(ws, 25, col, mg, fmt=_FMT_PCT)

    # ── Row 26: EBITDA % growth ────────────────────────────────────────────
    _lbl(ws, 26, 4, "  EBITDA % growth")
    ws.cell(row=26, column=9, value="—")
    for j in range(1, N):
        col = COLS[j]
        _cell(ws, 26, col,
              f"={_w(col)}24/{_w(col-1)}24-1",
              bg=_BG_CALC, fmt=_FMT_PCT)

    # ── Row 28: D&A ────────────────────────────────────────────────────────
    _lbl(ws, 28, 4, "D&A")
    for col in COLS:
        _cell(ws, 28, col, f"=-{_w(col)}21*$F$10",
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 29: EBIT ───────────────────────────────────────────────────────
    _lbl(ws, 29, 4, "EBIT")
    for col in COLS:
        _cell(ws, 29, col, f"={_w(col)}24+{_w(col)}28",
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 30: EBIT margin ────────────────────────────────────────────────
    _lbl(ws, 30, 4, "  EBIT margin")
    for col in COLS:
        _cell(ws, 30, col, f"={_w(col)}29/{_w(col)}21",
              bg=_BG_CALC, fmt=_FMT_PCT)

    # ── Row 32: Tax ────────────────────────────────────────────────────────
    _lbl(ws, 32, 4, "Tax")
    for col in COLS:
        _cell(ws, 32, col, f"=-{_w(col)}29*$F$13",
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 34: EBIAT ──────────────────────────────────────────────────────
    _lbl(ws, 34, 4, "EBIAT  (EBIT after tax)")
    for col in COLS:
        _cell(ws, 34, col, f"={_w(col)}29+{_w(col)}32",
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 36: D&A add-back ───────────────────────────────────────────────
    _lbl(ws, 36, 4, "Add: D&A")
    for col in COLS:
        _cell(ws, 36, col, f"=-{_w(col)}28",
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 38: Capex ──────────────────────────────────────────────────────
    _lbl(ws, 38, 4, "Capex")
    for col in COLS:
        _cell(ws, 38, col, f"=-{_w(col)}21*$F$11",
              bg=_BG_CALC, fmt=_FMT_NUM)

    _lbl(ws, 39, 4, "  % of revenue")
    for col in COLS:
        _cell(ws, 39, col, f"={_w(col)}38/{_w(col)}21",
              bg=_BG_CALC, fmt=_FMT_PCT)

    # ── Row 41: Change in NWC ──────────────────────────────────────────────
    _lbl(ws, 41, 4, "Change in NWC")
    for col in COLS:
        _cell(ws, 41, col, f"=-{_w(col)}21*$F$12",
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 43: Unlevered FCF ──────────────────────────────────────────────
    _lbl(ws, 43, 4, "Unlevered FCF")
    for col in COLS:
        _cell(ws, 43, col,
              f"={_w(col)}34+{_w(col)}36+{_w(col)}38+{_w(col)}41",
              bg=_BG_CALC, fmt=_FMT_NUM)

    # ── Row 45: Timing ─────────────────────────────────────────────────────
    _lbl(ws, 45, 4, "Cashflow timing (years)")
    for j, col in enumerate(COLS):
        _cell(ws, 45, col, j, fmt="0")

    # ── Row 46: Discount factor 1/(1+WACC)^t ──────────────────────────────
    _lbl(ws, 46, 4, "Discount factor  1/(1+WACC)^t")
    for j, col in enumerate(COLS):
        if j == 0:
            _cell(ws, 46, col, 1.0, fmt=_FMT_PRICE)
        else:
            _cell(ws, 46, col,
                  f"=1/(1+$F$9)^{_w(col)}45",
                  bg=_BG_CALC, fmt=_FMT_PRICE)

    # ── Row 48-50: Terminal Value ──────────────────────────────────────────
    _hdr(ws, 48, 4, "Terminal Value  (written in final projection year col S)",
         end_col=19, bg=_BG_SECT, sz=9)

    _lbl(ws, 49, 4, "TV — Perpetuity growth method")
    _out(ws, 49, 19,
         "=S43*(1+$F$7)/($F$9-$F$7)",
         fmt=_FMT_NUM)

    _lbl(ws, 50, 4, "TV — Exit EBITDA multiple")
    _out(ws, 50, 19,
         "=S24*$F$8",
         fmt=_FMT_NUM)

    ws.cell(row=48, column=20,
            value="TV uses FCF / EBITDA of final year (Y+10 = col S)"
            ).font = _f(8, False, _FG_G)


# ══════════════════════════════════════════════════════════════════════════
# Sheet 3: DCF output
# inp_name = exact title of the corresponding "DCF input" sheet
# ══════════════════════════════════════════════════════════════════════════


# ── Chart constants ────────────────────────────────────────────────────────
_CHART_FILL   = "BDD7EE"   # light blue bars (matches screenshot)
_CHART_BORDER = "2E75B6"   # darker blue border
_CHART_W      = 16         # chart width in cm
_CHART_H      = 10         # chart height in cm


def _add_fcf_chart(ws, fund: Optional[dict]) -> None:
    """
    Add a single Cash Flow bar chart to the DCF output sheet, placed below
    the sensitivity table (anchored at row 38).

    Shows Operating Cash Flow and Free Cash Flow side by side per year
    (up to 4 years, oldest→newest), matching the screenshot style:
      - Light-blue bars, value labels above each bar with number format
      - No axis titles, hidden Y-axis (labels carry the values)
      - No legend when only one series; legend shown for two series

    Data is written into hidden helper rows (rows 38-41) within the same
    sheet so the chart has a proper Reference source — not hardcoded values.
    """
    if not fund:
        return

    cf = fund.get("cashflow", {})
    if not cf:
        return

    # Collect years with at least one cash flow metric, oldest→newest
    years = sorted(cf.keys())[-4:]   # up to 4 years
    if not years:
        return

    ANCHOR_ROW   = 38   # chart placed here (row after legend note at row 35)
    HELPER_ROW_Y = 40   # year labels
    HELPER_ROW_O = 41   # Operating CF
    HELPER_ROW_F = 42   # Free CF
    DATA_START_COL = 3  # column C

    # ── Write hidden helper rows (white font = invisible to reader) ────────
    for j, yr in enumerate(years):
        col = DATA_START_COL + j   # C, D, E, F

        # Year label
        c = ws.cell(row=HELPER_ROW_Y, column=col, value=int(yr))
        c.font = Font(name="Calibri", size=8, color="FFFFFF")

        # Operating CF
        ocf_raw = cf.get(yr, {}).get("Operating Cash Flow") or 0
        ocf_val = round(float(ocf_raw) / 1e6, 1) if ocf_raw else 0
        c2 = ws.cell(row=HELPER_ROW_O, column=col, value=ocf_val)
        c2.font = Font(name="Calibri", size=8, color="FFFFFF")

        # Free CF
        fcf_raw = cf.get(yr, {}).get("Free Cash Flow") or 0
        fcf_val = round(float(fcf_raw) / 1e6, 1) if fcf_raw else 0
        c3 = ws.cell(row=HELPER_ROW_F, column=col, value=fcf_val)
        c3.font = Font(name="Calibri", size=8, color="FFFFFF")

    n = len(years)
    min_col = DATA_START_COL
    max_col = DATA_START_COL + n - 1

    # ── Build chart ────────────────────────────────────────────────────────
    chart = BarChart()
    chart.type     = "col"
    chart.grouping = "clustered"
    chart.title    = "Cash Flow ($m)"
    chart.width    = _CHART_W
    chart.height   = _CHART_H

    # Clean axes: no titles, hide Y axis (values shown as labels)
    chart.x_axis.title = None
    chart.y_axis.title = None
    chart.y_axis.delete = True

    # Category reference (year labels)
    cats = Reference(ws,
                     min_col=min_col, max_col=max_col,
                     min_row=HELPER_ROW_Y, max_row=HELPER_ROW_Y)

    # Add Operating CF series
    ocf_data = Reference(ws,
                         min_col=min_col, max_col=max_col,
                         min_row=HELPER_ROW_O, max_row=HELPER_ROW_O)
    chart.add_data(ocf_data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    s_ocf = chart.series[0]
    s_ocf.title = openpyxl.chart.series.SeriesLabel(v="Operating CF")
    s_ocf.graphicalProperties.solidFill      = "A9C4E4"   # slightly darker blue
    s_ocf.graphicalProperties.line.solidFill = _CHART_BORDER
    s_ocf.graphicalProperties.line.width     = 6350

    dl_ocf = DataLabelList()
    dl_ocf.showVal       = True
    dl_ocf.showCatName   = False
    dl_ocf.showSerName   = False
    dl_ocf.showLegendKey = False
    dl_ocf.position      = "outEnd"
    dl_ocf.numFmt        = "#,##0.0"
    s_ocf.dLbls = dl_ocf

    # Add Free CF series
    fcf_data = Reference(ws,
                         min_col=min_col, max_col=max_col,
                         min_row=HELPER_ROW_F, max_row=HELPER_ROW_F)
    chart.add_data(fcf_data, from_rows=True, titles_from_data=False)
    s_fcf = chart.series[1]
    s_fcf.title = openpyxl.chart.series.SeriesLabel(v="Free CF")
    s_fcf.graphicalProperties.solidFill      = _CHART_FILL   # lighter blue
    s_fcf.graphicalProperties.line.solidFill = _CHART_BORDER
    s_fcf.graphicalProperties.line.width     = 6350

    dl_fcf = DataLabelList()
    dl_fcf.showVal       = True
    dl_fcf.showCatName   = False
    dl_fcf.showSerName   = False
    dl_fcf.showLegendKey = False
    dl_fcf.position      = "outEnd"
    dl_fcf.numFmt        = "#,##0.0"
    s_fcf.dLbls = dl_fcf

    anchor = f"C{ANCHOR_ROW}"
    ws.add_chart(chart, anchor)


def _build_dcf_output(wb, rec: dict, fund: Optional[dict],
                      suffix: str = "", inp_name: str = "DCF input"):
    ws = wb.create_sheet(f"DCF output{suffix}")
    for col, w in [
        ("A",3),("B",3),("C",30),("D",3),("E",3),("F",14),
        ("G",12),("H",3),("I",3),("J",14),("K",3),
        ("L",12),("M",12),("N",12),("O",12),("P",12),
    ]:
        ws.column_dimensions[col].width = w

    name   = rec.get("name")   or rec.get("ticker","X")
    cur    = rec.get("currency") or "USD"
    region = rec.get("region",  "US")
    price  = rec.get("price")

    # Net debt
    debt_m  = _get_fund(fund, "balance", "Total Debt", 0.0) or 0.0
    cash_m  = _get_fund(fund, "balance", "Cash And Cash Equivalents", 0.0) or 0.0
    net_debt = round(debt_m - cash_m, 1)

    # Shares
    shares = None
    if fund:
        sh = _get_fund(fund, "balance", "Share Issued")
        if sh and sh > 0:
            shares = round(sh, 2)
    if not shares and rec.get("market_cap") and rec.get("price"):
        try:
            shares = round(float(rec["market_cap"]) / float(rec["price"]) / 1e6, 2)
        except (TypeError, ValueError, ZeroDivisionError):
            shares = 100.0
    shares = shares or 100.0

    # Title
    ws.merge_cells("A1:P1")
    c = ws.cell(row=1, column=1, value=f"DCF Output — {name}  [{cur}m]")
    c.font = Font(name="Calibri", size=12, bold=True, color=_FG_W)
    c.fill = _fill(_BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Note row numbers: input sheet changed (UFCF=row43, DF=row46, TV=rows49/50)
    ucf_row = 43
    df_row  = 46
    tv_pgm  = 49
    tv_mult = 50

    # ── PGM ───────────────────────────────────────────────────────────────
    _hdr(ws, 3, 3, "Net Present Value — Perpetuity Growth Method",
         end_col=9, bg=_BG_HEAD)
    _hdr(ws, 4, 6, "$m",    bg=_BG_HEAD, sz=9)
    _hdr(ws, 4, 7, "% NPV", bg=_BG_HEAD, sz=9)

    _lbl(ws, 5, 3, "PV of FCFs")
    _out(ws, 5, 6,
         f"=SUMPRODUCT('{inp_name}'!J{ucf_row}:S{ucf_row},'{inp_name}'!J{df_row}:S{df_row})",
         fmt=_FMT_NUM)
    _cell(ws, 5, 7, "=F5/F$9", bg=_BG_CALC, fmt=_FMT_PCT)

    _lbl(ws, 6, 3, "PV of Terminal Value (PGM)")
    _out(ws, 6, 6,
         f"='{inp_name}'!S{tv_pgm}*'{inp_name}'!S{df_row}",
         fmt=_FMT_NUM)
    _cell(ws, 6, 7, "=F6/F$9", bg=_BG_CALC, fmt=_FMT_PCT)

    _lbl(ws, 7, 3, "Implied Firm Value (EV)")
    _out(ws, 7, 6, "=SUM(F5:F6)", fmt=_FMT_NUM)
    _cell(ws, 7, 7, "=F7/F$7", bg=_BG_CALC, fmt=_FMT_PCT)

    _lbl(ws, 8, 3, "Less: Net Debt  (Debt − Cash)")
    _inp(ws, 8, 6, net_debt, fmt=_FMT_NUM)

    _lbl(ws, 9, 3, "Implied Equity Value")
    _out(ws, 9, 6, "=F7-F8", fmt=_FMT_NUM)

    _lbl(ws, 10, 3, "Shares outstanding (m)")
    _inp(ws, 10, 6, shares)

    _lbl(ws, 11, 3, "DCF Price per share  (PGM)")
    _out(ws, 11, 6, "=F9/F10", fmt=_FMT_PRICE)

    # ── EV/EBITDA multiple ────────────────────────────────────────────────
    _hdr(ws, 13, 3, "Net Present Value — Exit EBITDA Multiple Method",
         end_col=9, bg=_BG_HEAD)
    _hdr(ws, 14, 6, "$m",    bg=_BG_HEAD, sz=9)
    _hdr(ws, 14, 7, "% NPV", bg=_BG_HEAD, sz=9)

    _lbl(ws, 15, 3, "PV of FCFs")
    _out(ws, 15, 6,
         f"=SUMPRODUCT('{inp_name}'!J{ucf_row}:S{ucf_row},'{inp_name}'!J{df_row}:S{df_row})",
         fmt=_FMT_NUM)
    _cell(ws, 15, 7, "=F15/F$17", bg=_BG_CALC, fmt=_FMT_PCT)

    _lbl(ws, 16, 3, "PV of Terminal Value (EV/EBITDA)")
    _out(ws, 16, 6,
         f"='{inp_name}'!S{tv_mult}*'{inp_name}'!S{df_row}",
         fmt=_FMT_NUM)
    _cell(ws, 16, 7, "=F16/F$17", bg=_BG_CALC, fmt=_FMT_PCT)

    _lbl(ws, 17, 3, "Implied Firm Value (EV)")
    _out(ws, 17, 6, "=SUM(F15:F16)", fmt=_FMT_NUM)
    _cell(ws, 17, 7, "=F17/F$17", bg=_BG_CALC, fmt=_FMT_PCT)

    _lbl(ws, 18, 3, "Less: Net Debt")
    _out(ws, 18, 6, "=F8", fmt=_FMT_NUM)

    _lbl(ws, 19, 3, "Implied Equity Value")
    _out(ws, 19, 6, "=F17-F18", fmt=_FMT_NUM)

    _lbl(ws, 20, 3, "DCF Price per share  (EV/EBITDA)")
    _out(ws, 20, 6, "=F19/F10", fmt=_FMT_PRICE)

    # ── vs market ─────────────────────────────────────────────────────────
    if price is not None:
        _hdr(ws, 22, 3, "vs. Current Market Price", end_col=9,
             bg=_BG_SECT, sz=9)
        _lbl(ws, 23, 3, "Current market price")
        _inp(ws, 23, 6, round(float(price), 2), fmt=_FMT_PRICE)
        _lbl(ws, 24, 3, "Upside / downside  (PGM)")
        _out(ws, 24, 6, "=(F11-F23)/F23", fmt=_FMT_PCT)
        _lbl(ws, 25, 3, "Upside / downside  (EV/EBITDA)")
        _out(ws, 25, 6, "=(F20-F23)/F23", fmt=_FMT_PCT)

    # ── Sensitivity ───────────────────────────────────────────────────────
    _hdr(ws, 27, 3, "Sensitivity — Implied Equity Value ($m)",
         end_col=9, bg=_BG_HEAD)

    wacc_base = {"KZ":0.14,"Emerging":0.12,"Europe":0.09,"Asia":0.10
                 }.get(region, 0.094)
    tgr_base  = {"KZ":0.04,"Emerging":0.035}.get(region, 0.025)

    wacc_steps = [-0.02, -0.01, 0.0, 0.01, 0.02]
    tgr_steps  = [-0.01, -0.005, 0.0, 0.005, 0.01]

    lc = ws.cell(row=28, column=3, value="WACC \\ TGR")
    lc.font = _f(9, True, _FG_D); lc.border = _bd()
    for j, dt in enumerate(tgr_steps):
        c2 = ws.cell(row=28, column=4+j, value=f"{(tgr_base+dt)*100:.1f}%")
        c2.font = _f(9, True, _FG_W)
        c2.fill = _fill(_BG_HEAD); c2.border = _bd()
        c2.alignment = Alignment(horizontal="center", vertical="center")

    rev_m    = _get_fund(fund, "income", "Total Revenue", 0.0) or 0.0
    ebitda_m = _get_fund(fund, "income", "EBITDA", None)
    emg      = _safe_ratio(ebitda_m, rev_m, default=0.20, min_val=0.0, max_val=0.90)

    for i, dw in enumerate(wacc_steps):
        row = 29 + i
        w_v = wacc_base + dw
        wc  = ws.cell(row=row, column=3, value=f"{w_v*100:.1f}%")
        wc.font = _f(9, True, _FG_W)
        wc.fill = _fill(_BG_HEAD); wc.border = _bd()
        for j, dt in enumerate(tgr_steps):
            t_v = tgr_base + dt
            try:
                eff_w      = max(w_v, t_v + 0.001)
                rev_y10    = rev_m * (1.10 ** 10)
                ebitda_y10 = rev_y10 * emg
                # Approximate UFCF in final year
                ufcf_y10   = ebitda_y10 * (1 - 0.20) * 0.80
                tv         = ufcf_y10 * (1 + t_v) / (eff_w - t_v)
                pv_tv      = tv / (1 + eff_w) ** 10
                pv_fcf     = sum(
                    rev_m * (1.08 ** k) * emg * (1-0.20) * 0.80
                    / (1 + eff_w) ** k
                    for k in range(1, 11)
                )
                eq = pv_tv + pv_fcf - net_debt
                is_base = abs(dw) < 0.001 and abs(dt) < 0.001
                _cell(ws, row, 4+j, round(eq, 0),
                      bg=_BG_OUT if is_base else _BG_CALC, fmt=_FMT_NUM)
            except Exception:
                _cell(ws, row, 4+j, "N/A")

    ws.cell(row=35, column=3,
            value="Blue = editable input  |  Orange = calculated output  "
                  "|  Sensitivity is approximate — adjust DCF input for exact values"
            ).font = _f(8, False, _FG_G)

    # ── Cash Flow chart ────────────────────────────────────────────────────
    _add_fcf_chart(ws, fund)


# ══════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════

def build_dcf_multi(records: list, fund_map: Optional[dict] = None) -> bytes:
    """
    Build DCF Excel workbook.
    Single ticker  → 3 sheets with no suffix.
    Multiple tickers → 3 sheets per ticker with ' TICKER' suffix.
    Cross-sheet formulas always reference the correct sheet name.
    """
    fund_map = fund_map or {}
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    multi = len(records) > 1

    for rec in records:
        ticker   = (rec.get("ticker") or "X").replace("/","-")[:8]
        fund     = fund_map.get(rec.get("ticker",""))
        suffix   = f" {ticker}" if multi else ""
        inp_name = f"DCF input{suffix}"

        _build_forecasts(wb, rec, fund, suffix)
        _build_dcf_input(wb, rec, fund, suffix)
        _build_dcf_output(wb, rec, fund, suffix, inp_name=inp_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
