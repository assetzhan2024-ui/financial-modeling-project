"""
export/excel.py  — строго по шаблону пример_для_проекта.xlsx
"""
import io
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("pip install openpyxl")

# ── Палитра ────────────────────────────────────────────────────────────────
_BG_DARK  = "1F2D3D"
_BG_SECT  = "2E4057"
_BG_LBL   = "EBF5FB"
_BG_VAL   = "FFFFFF"
_BG_ALT   = "F4F6F7"
_BG_YEAR  = "154360"
_FG_W     = "FFFFFF"
_FG_D     = "1C2833"
_FG_G     = "717D7E"
_FG_GRN   = "1E8449"
_FG_RED   = "C0392B"
_BRD      = "BFC9CA"

def _bd():
    s = Side(style="thin", color=_BRD)
    return Border(left=s, right=s, top=s, bottom=s)
def _fill(h): return PatternFill("solid", fgColor=h)
def _fnt(sz=10, bold=False, color=_FG_D):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def _fB(v):
    if v is None: return "—"
    try: v = float(v)
    except: return str(v)
    neg = v < 0; a = abs(v)
    if a == 0: return "0"
    if a >= 1e12: s = f"{a/1e12:.2f}T"
    elif a >= 1e9: s = f"{a/1e9:.2f}B"
    elif a >= 1e6: s = f"{a/1e6:.2f}M"
    elif a >= 1e3: s = f"{a/1e3:.1f}K"
    else: s = f"{a:.2f}"
    return f"({s})" if neg else s

def _fN(v, d=2):
    if v is None: return "—"
    try: return round(float(v), d)
    except: return "—"

def _fP(v):
    if v is None: return "—"
    try: return f"{float(v):.1f}%"
    except: return "—"

def _rc(rat):
    return {"ideal":_FG_GRN,"good":_FG_GRN,"warn":_FG_RED}.get(rat or "", _FG_D)

# ── Запись ячейки-метки ────────────────────────────────────────────────────
def _L(ws, row, col, text, bg=_BG_LBL):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _fnt(10, False, _FG_G); c.fill = _fill(bg)
    c.border = _bd(); c.alignment = Alignment(horizontal="left", vertical="center")

def _V(ws, row, col, v, color=_FG_D, bg=_BG_VAL, bold=False):
    c = ws.cell(row=row, column=col, value=v)
    c.font = Font(name="Calibri", size=10, bold=bold, color=color)
    c.fill = _fill(bg); c.border = _bd()
    c.alignment = Alignment(horizontal="right", vertical="center")

def _S(ws, row, col, text, end_col=None, bg=_BG_SECT):
    if end_col and end_col > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=10, bold=True, color=_FG_W)
    c.fill = _fill(bg); c.border = _bd()
    c.alignment = Alignment(horizontal="left", vertical="center")

def _fin(ws, row, col_h, label, src, key, years, col_start=9):
    alt = row % 2 == 0; bg = _BG_ALT if alt else _BG_VAL
    _L(ws, row, col_h, label, bg=bg)
    for j, yr in enumerate(years):
        raw = src.get(yr, {}).get(key)
        col = col_start + j
        fg = (_FG_GRN if raw is not None and float(raw) > 0
              else (_FG_RED if raw is not None else _FG_G))
        _V(ws, row, col, _fB(raw), color=fg, bg=bg)


# ══════════════════════════════════════════════════════════════════════════
# Лист тикера — строго по шаблону
# ══════════════════════════════════════════════════════════════════════════
def _ticker_sheet(wb, rec: dict, fund: dict | None):
    ticker = rec.get("ticker","X")
    name   = rec.get("name") or ticker
    ws     = wb.create_sheet(title=ticker[:31])

    # Ширины столбцов (по шаблону)
    for col,w in [("A",12),("B",16),("C",2),("D",16),("E",14),
                  ("F",2),("G",2),("H",22),("I",14),("J",14),("K",14),("L",14)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22

    # ── Row 1: шапка ─────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value=f"{name}  ({ticker})")
    c.font = Font(name="Calibri", size=12, bold=True, color=_FG_W)
    c.fill = _fill(_BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Годы ─────────────────────────────────────────────────────────────
    if fund:
        years = sorted({*fund.get("income",{}).keys(),
                        *fund.get("balance",{}).keys(),
                        *fund.get("cashflow",{}).keys()},
                       reverse=True)[:4]
    else:
        years = []

    # I2, J2, K2, L2 — заголовки годов
    for j, yr in enumerate(years):
        c2 = ws.cell(row=2, column=9+j, value=int(yr))
        c2.font = Font(name="Calibri", size=11, bold=True, color=_FG_W)
        c2.fill = _fill(_BG_YEAR); c2.border = _bd()
        c2.alignment = Alignment(horizontal="right", vertical="center")

    # ── Левый блок ───────────────────────────────────────────────────────
    r   = rec
    rats= r.get("ratings", {})
    cur = r.get("currency") or "USD"
    iUSD= cur == "USD"

    # A3: market cap / B3: значение
    mc  = r.get("market_cap_usd") or r.get("market_cap")
    mc_txt = f"${_fB(mc)}"
    if not iUSD and r.get("market_cap") and mc:
        mc_txt += f"  ({_fB(r['market_cap'])} {cur})"
    _L(ws, 3, 1, "market cap")    # A3
    _V(ws, 3, 2, mc_txt)          # B3

    # D3: price / E3: значение
    p_usd = r.get("price_usd") or r.get("price")
    p_txt = f"${_fN(p_usd)}" if p_usd is not None else "—"
    if not iUSD and r.get("price") and p_usd:
        p_txt += f"  ({_fN(r['price'])} {cur})"
    _L(ws, 3, 4, "price")         # D3
    _V(ws, 3, 5, p_txt)           # E3

    # D4: range52 weeks / E4
    lo = r.get("week52_low"); hi = r.get("week52_high")
    _L(ws, 4, 4, "range52 weeks")                          # D4
    _V(ws, 4, 5, f"{_fN(lo)} – {_fN(hi)}" if lo else "—") # E4

    # D7..D14: мультипликаторы / E7..E14
    def lv(row, lbl, val, rkey=None):
        _L(ws, row, 4, lbl)
        _V(ws, row, 5, val, color=_rc(rats.get(rkey)) if rkey else _FG_D)

    lv(7,  "p/e",             _fN(r.get("pe_ratio")),         "pe_ratio")
    lv(8,  "p/b",             _fN(r.get("pb_ratio")),         "pb_ratio")
    lv(9,  "p/s",             _fN(r.get("ps_ratio")),         "ps_ratio")
    lv(10, "ev/ebitda",       _fN(r.get("ev_ebitda")),        "ev_ebitda")
    lv(11, "roe%",            _fP(r.get("roe_pct")),          "roe_pct")
    lv(12, "d/e",             _fN(r.get("de_ratio")),         "de_ratio")
    lv(13, "net debt/ebitda", _fN(r.get("net_debt_ebitda")),  "net_debt_ebitda")
    lv(14, "eps",             _fN(r.get("eps_trailing_usd") or r.get("eps_trailing")))

    # ── Правый блок: финансовая отчётность ───────────────────────────────
    inc = (fund or {}).get("income",  {})
    bal = (fund or {}).get("balance", {})
    cf  = (fund or {}).get("cashflow",{})
    end = 8 + max(len(years), 1)

    # Income statement — H3..H8
    _S(ws, 3, 8, "income statement", end_col=end)   # H3
    _fin(ws, 4,  8, "revenue",        inc, "Total Revenue",       years)
    _fin(ws, 5,  8, "gross profit",   inc, "Gross Profit",        years)
    _fin(ws, 6,  8, "operate income", inc, "Operating Income",    years)
    _fin(ws, 7,  8, "ebitda",         inc, "EBITDA",              years)
    _fin(ws, 8,  8, "net income",     inc, "Net Income",          years)

    # Balance sheet — H11..H16
    _S(ws, 11, 8, "balance sheet", end_col=end)     # H11
    _fin(ws, 12, 8, "assets",               bal, "Total Assets",                             years)
    _fin(ws, 13, 8, "liabilities",          bal, "Total Liabilities Net Minority Interest",  years)
    _fin(ws, 14, 8, "stockholders equity",  bal, "Stockholders Equity",                      years)
    _fin(ws, 15, 8, "debt",                 bal, "Total Debt",                               years)
    _fin(ws, 16, 8, "cash&cash equivalence",bal, "Cash And Cash Equivalents",                years)

    # Cash flow — H18..H22
    _S(ws, 18, 8, "cash flow statement", end_col=end)  # H18
    _fin(ws, 19, 8, "operating cash flow", cf, "Operating Cash Flow",  years)
    _fin(ws, 20, 8, "free cash flow",      cf, "Free Cash Flow",       years)
    _fin(ws, 21, 8, "Capex",              cf, "Capital Expenditure",   years)
    _fin(ws, 22, 8, "investing cash flow", cf, "Investing Cash Flow",  years)

    if not fund or not years:
        ws.cell(row=4, column=9,
                value="Нет данных — откройте карточку тикера для загрузки"
                ).font = _fnt(9, False, _FG_G)


# ══════════════════════════════════════════════════════════════════════════
# Лист Screener — сводная таблица (без EPS forward)
# ══════════════════════════════════════════════════════════════════════════
def _screener_sheet(wb, records: list[dict]):
    ws = wb.create_sheet(title="Screener", index=0)

    HDR = [
        ("Тикер",      14), ("Название",   26), ("Регион",    10), ("Сектор",    18),
        ("Валюта",      8), ("Score",       8),
        ("Цена (USD)", 13), ("Market Cap (USD)", 18),
        ("P/E",         8), ("P/B",         8), ("P/S",        8),
        ("EV/EBITDA",  11), ("EV/Revenue",  11), ("D/E",        8), ("ND/EBITDA", 11),
        ("ROE %",       9), ("ROA %",       9),
        ("EBITDA (USD)",15),("Net Inc (USD)",15),("FCF (USD)",  12),
        ("EPS",        12), ("52W Low",    10), ("52W High",   10),
    ]
    ws.row_dimensions[1].height = 28
    for i,(label,w) in enumerate(HDR, 1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        c.fill = _fill(_BG_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _bd()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    for ri, b in enumerate(records, 2):
        bg   = _BG_ALT if ri%2==0 else _BG_VAL
        rats = b.get("ratings", {})
        ws.row_dimensions[ri].height = 16

        def C(col, txt, rat=None, bold=False):
            color = _rc(rats.get(rat)) if rat else _FG_D
            c2 = ws.cell(row=ri, column=col, value=txt)
            c2.font = Font(name="Calibri", size=9, bold=bold, color=color)
            c2.fill = _fill(bg)
            c2.alignment = Alignment(
                horizontal="left" if col<=4 else "right", vertical="center")
            c2.border = _bd()

        C(1,  b.get("ticker",""),                             bold=True)
        C(2,  b.get("name","—"))
        C(3,  b.get("region","—"))
        C(4,  b.get("sector","—"))
        C(5,  b.get("currency","USD"))
        C(6,  f"{b.get('score_pct',0)}%",                    bold=True)
        C(7,  f"${_fN(b.get('price_usd') or b.get('price'))}")
        C(8,  f"${_fB(b.get('market_cap_usd') or b.get('market_cap'))}")
        C(9,  _fN(b.get("pe_ratio")),                        "pe_ratio")
        C(10, _fN(b.get("pb_ratio")),                        "pb_ratio")
        C(11, _fN(b.get("ps_ratio")),                        "ps_ratio")
        C(12, _fN(b.get("ev_ebitda")),                       "ev_ebitda")
        C(13, _fN(b.get("ev_revenue")))
        C(14, _fN(b.get("de_ratio")),                        "de_ratio")
        C(15, _fN(b.get("net_debt_ebitda")),                 "net_debt_ebitda")
        C(16, _fP(b.get("roe_pct")),                         "roe_pct")
        C(17, _fP(b.get("roa_pct")))
        C(18, f"${_fB(b.get('ebitda_usd')     or b.get('ebitda'))}")
        C(19, f"${_fB(b.get('net_income_usd') or b.get('net_income'))}")
        C(20, f"${_fB(b.get('fcf_usd')        or b.get('fcf'))}")
        C(21, _fN(b.get("eps_trailing_usd")   or b.get("eps_trailing")))
        C(22, _fN(b.get("week52_low")))
        C(23, _fN(b.get("week52_high")))


# ══════════════════════════════════════════════════════════════════════════
# Публичный API
# ══════════════════════════════════════════════════════════════════════════
def build_excel(records: list[dict], fund_map: dict | None = None) -> bytes:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _screener_sheet(wb, records)
<<<<<<< HEAD
    # CCA — только если несколько тикеров
    if len(records) >= 2:
        _cca_sheet(wb, records)
=======
>>>>>>> 48fe0a82dc8de8e4b1571d8dc69fbb5300300ae0
    for rec in records:
        fd = (fund_map or {}).get(rec.get("ticker",""))
        _ticker_sheet(wb, rec, fd)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
<<<<<<< HEAD


# ══════════════════════════════════════════════════════════════════════════
# Лист CCA — Comparable Company Analysis
# ══════════════════════════════════════════════════════════════════════════

# Метрики для CCA — (display_label, field_key, higher_is_better)
_CCA_METRICS = [
    ("P/E",             "pe_ratio",         False),
    ("P/B",             "pb_ratio",         False),
    ("P/S",             "ps_ratio",         False),
    ("EV/EBITDA",       "ev_ebitda",        False),
    ("EV/Revenue",      "ev_revenue",       False),
    ("D/E",             "de_ratio",         False),
    ("ND/EBITDA",       "net_debt_ebitda",  False),
    ("ROE %",           "roe_pct",          True),
    ("ROA %",           "roa_pct",          True),
    ("EPS (trail.)",    "eps_trailing",     True),
]

_BG_BEST   = "D5F5E3"   # светло-зелёный — лучшее значение в строке
_BG_WORST  = "FADBD8"   # светло-красный — худшее
_BG_AVG    = "EBF5FB"   # голубой — строка avg/median
_FG_BEST   = "1E8449"
_FG_WORST  = "C0392B"
_FG_AVG    = "1A5276"


def _cca_sheet(wb, records: list[dict]):
    """
    Создать лист CCA (Comparable Company Analysis).

    Структура:
      Строка 1:  шапка
      Строка 2:  заголовки компаний (тикер + название)
      Строки 3…: одна строка на метрику, значения по компаниям
      После всех метрик — блок Avg / Median / Discount-Premium
    """
    if len(records) < 2:
        return   # CCA не имеет смысла для одной компании

    ws = wb.create_sheet(title="CCA")

    N = len(records)  # количество тикеров

    # ── Ширины столбцов ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18   # метрика
    ws.column_dimensions["B"].width = 14   # avg
    ws.column_dimensions["C"].width = 14   # median
    for i in range(N):
        col = get_column_letter(4 + i)     # D, E, F, … — по одной на тикер
        ws.column_dimensions[col].width = 14

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # ── Row 1: шапка ─────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + N)
    c = ws.cell(row=1, column=1, value="Comparable Company Analysis (CCA)")
    c.font      = Font(name="Calibri", size=13, bold=True, color=_FG_W)
    c.fill      = _fill(_BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 2: колонки Avg / Median / [тикеры] ──────────────────────────
    def hdr2(col, text, bg=_BG_SECT):
        c2 = ws.cell(row=2, column=col, value=text)
        c2.font      = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        c2.fill      = _fill(bg)
        c2.border    = _bd()
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hdr2(1, "Метрика",  _BG_DARK)
    hdr2(2, "Среднее",  _BG_AVG.replace("EBF5FB","1A5276"))   # тёмно-синий
    hdr2(3, "Медиана",  _BG_AVG.replace("EBF5FB","154360"))

    for i, rec in enumerate(records):
        label = f"{rec.get('ticker','?')}\n{(rec.get('name') or '')[:18]}"
        hdr2(4 + i, label, _BG_YEAR)

    # ── Строки метрик ────────────────────────────────────────────────────
    for row_idx, (label, field, higher_is_better) in enumerate(_CCA_METRICS, start=3):
        alt = row_idx % 2 == 0
        bg  = _BG_ALT if alt else _BG_VAL

        # Значения по всем тикерам
        vals = []
        for rec in records:
            v = rec.get(field)
            try:   vals.append(float(v) if v is not None else None)
            except: vals.append(None)

        valid_vals = [v for v in vals if v is not None]

        # Avg и Median
        avg = sum(valid_vals) / len(valid_vals) if valid_vals else None
        med = sorted(valid_vals)[len(valid_vals) // 2] if valid_vals else None

        # Лучший/худший индекс (с учётом направления)
        best_i = worst_i = -1
        if valid_vals:
            best_val  = max(valid_vals) if higher_is_better else min(valid_vals)
            worst_val = min(valid_vals) if higher_is_better else max(valid_vals)
            for i, v in enumerate(vals):
                if v == best_val  and best_i  < 0: best_i  = i
                if v == worst_val and worst_i < 0: worst_i = i

        # Метрика-метка
        lc = ws.cell(row=row_idx, column=1, value=label)
        lc.font      = _fnt(10, False, _FG_G)
        lc.fill      = _fill(_BG_LBL)
        lc.border    = _bd()
        lc.alignment = Alignment(horizontal="left", vertical="center")

        # Avg
        avg_txt = (_fP(avg) if field in ("roe_pct","roa_pct") else _fN(avg)) if avg is not None else "—"
        ac = ws.cell(row=row_idx, column=2, value=avg_txt)
        ac.font      = Font(name="Calibri", size=10, bold=True, color=_FG_AVG)
        ac.fill      = _fill(_BG_AVG)
        ac.border    = _bd()
        ac.alignment = Alignment(horizontal="right", vertical="center")

        # Median
        med_txt = (_fP(med) if field in ("roe_pct","roa_pct") else _fN(med)) if med is not None else "—"
        mc = ws.cell(row=row_idx, column=3, value=med_txt)
        mc.font      = Font(name="Calibri", size=10, bold=True, color=_FG_AVG)
        mc.fill      = _fill(_BG_AVG)
        mc.border    = _bd()
        mc.alignment = Alignment(horizontal="right", vertical="center")

        # Значения по тикерам
        for i, v in enumerate(vals):
            is_best  = (i == best_i)
            is_worst = (i == worst_i and i != best_i)

            if field in ("roe_pct", "roa_pct"):
                txt = _fP(v) if v is not None else "—"
            else:
                txt = _fN(v) if v is not None else "—"

            cell_bg = _BG_BEST if is_best else (_BG_WORST if is_worst else bg)
            cell_fg = _FG_BEST if is_best else (_FG_WORST if is_worst else _FG_D)
            bold    = is_best or is_worst

            vc = ws.cell(row=row_idx, column=4 + i, value=txt)
            vc.font      = Font(name="Calibri", size=10, bold=bold, color=cell_fg)
            vc.fill      = _fill(cell_bg)
            vc.border    = _bd()
            vc.alignment = Alignment(horizontal="right", vertical="center")

    # ── Разделитель ──────────────────────────────────────────────────────
    sep_row = 3 + len(_CCA_METRICS) + 1
    ws.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=3 + N)
    sh = ws.cell(row=sep_row, column=1,
                 value="Discount / Premium к медиане группы (медиана = 0%)")
    sh.font      = Font(name="Calibri", size=9, bold=True, color=_FG_W)
    sh.fill      = _fill(_BG_SECT)
    sh.alignment = Alignment(horizontal="left", vertical="center")

    # ── Строки Discount/Premium ───────────────────────────────────────────
    DISC_METRICS = [
        ("P/E",         "pe_ratio",   False),
        ("EV/EBITDA",   "ev_ebitda",  False),
        ("EV/Revenue",  "ev_revenue", False),
        ("P/B",         "pb_ratio",   False),
    ]
    for di, (label, field, _) in enumerate(DISC_METRICS):
        dr = sep_row + 1 + di
        alt = dr % 2 == 0
        bg  = _BG_ALT if alt else _BG_VAL

        vals = []
        for rec in records:
            v = rec.get(field)
            try:   vals.append(float(v) if v is not None else None)
            except: vals.append(None)

        valid_vals = [v for v in vals if v is not None]
        med = sorted(valid_vals)[len(valid_vals) // 2] if valid_vals else None

        lc2 = ws.cell(row=dr, column=1, value=f"{label} vs. med")
        lc2.font = _fnt(10, False, _FG_G); lc2.fill = _fill(_BG_LBL)
        lc2.border = _bd()

        # Avg/Med колонки пустые в discount блоке
        for col in [2, 3]:
            ec = ws.cell(row=dr, column=col, value="—")
            ec.fill = _fill(_BG_AVG); ec.border = _bd()
            ec.alignment = Alignment(horizontal="right", vertical="center")
            ec.font = _fnt(9, False, _FG_G)

        for i, v in enumerate(vals):
            if v is None or med is None or med == 0:
                txt = "—"; fg = _FG_G; cell_bg = bg; bold = False
            else:
                disc = (v - med) / abs(med)   # отрицательное = дешевле медианы
                pct  = disc * 100
                # Для мультипликаторов: дешевле медианы = хорошо (зелёный)
                positive_is_good = False   # для P/E, P/B ниже = лучше
                txt  = f"{'+' if pct >= 0 else ''}{pct:.1f}%"
                if (pct < -5 and not positive_is_good) or (pct > 5 and positive_is_good):
                    fg = _FG_BEST; cell_bg = _BG_BEST; bold = True   # дешевле
                elif (pct > 5 and not positive_is_good) or (pct < -5 and positive_is_good):
                    fg = _FG_WORST; cell_bg = _BG_WORST; bold = True  # дороже
                else:
                    fg = _FG_D; cell_bg = bg; bold = False

            vc = ws.cell(row=dr, column=4 + i, value=txt)
            vc.font      = Font(name="Calibri", size=10, bold=bold, color=fg)
            vc.fill      = _fill(cell_bg)
            vc.border    = _bd()
            vc.alignment = Alignment(horizontal="right", vertical="center")

    # ── Легенда ──────────────────────────────────────────────────────────
    leg_row = sep_row + 1 + len(DISC_METRICS) + 1
    ws.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=3 + N)
    lc3 = ws.cell(row=leg_row, column=1,
                  value="🟢 Зелёный = лучшее значение в строке  |  🔴 Красный = худшее  "
                        "|  Discount/Premium: отрицательный % = дешевле медианы группы")
    lc3.font = Font(name="Calibri", size=8, italic=True, color=_FG_G)
    ws.row_dimensions[leg_row].height = 14
=======
>>>>>>> 48fe0a82dc8de8e4b1571d8dc69fbb5300300ae0
