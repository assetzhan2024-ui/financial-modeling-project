"""
export/dcf.py
=============
DCF Valuation Excel — строго по шаблону dcf_hsbk.xlsx.

Структура листа:
  Row 1:  A1=income statement   E1..J1=годы (0..5)   L1=5
  Row 2:  A2=revenue            E2=base  F2..J2=формулы роста
  Row 3:  A3=%growth            F3..J3=ставки роста
  Row 6:  A6=ebitda             E6..J6=revenue*margin
  Row 7:  A7=% margin           E7..J7=margin values
  Row 9:  A9=d&a                E9..J9=formulas
  Row 10: A10=ebit              E10..J10=ebitda-d&a
  Row 13: A13=free cash flow statement
  Row 14: A14=ebitda
  Row 15: A15=tax
  Row 16: A16=capex
  Row 17: A17=change in net wc
  Row 18: A18=unlevered cash flow
  Row 21:  L21=perpetual grow rate  M21=value
  Row 22: A22=discount cash flow
  Row 23: A23=annual unlevered free cash flow
  Row 24: A24=terminal value
  Row 25: A25=total unlevered free cash flow
  Row 26: L26=capm  M26=value
  Row 27: L27=wacc  M27=value
  Row 28: A28=present value ufcf  D28=sum  F28..J28=PV formulas
  Row 29: A29=net pv  D29=NPV
  Row 32: A32=sum of fcff
  Row 33: A33=cash/cash equivalents
  Row 34: A34=total debt
  Row 35: A35=shares outstanding
  Row 36: A36=equity value
  Row 37: A37=dcf price per share

Публичный API:
    build_dcf(record, fund_data=None) -> bytes
    build_dcf_multi(records, fund_map=None) -> bytes   (несколько листов)
"""

import io
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("pip install openpyxl")


# ── Стили ──────────────────────────────────────────────────────────────────
_BG_DARK  = "1F2D3D"
_BG_SECT  = "2C3E50"
_BG_INPUT = "EBF5FB"   # голубой — редактируемые входные данные
_BG_CALC  = "FDFEFE"   # белый — расчётные ячейки
_BG_OUT   = "F9EBEA"   # розовый — итоговые значения
_BG_YEAR  = "154360"
_FG_W     = "FFFFFF"
_FG_D     = "1C2833"
_FG_G     = "717D7E"
_FG_GRN   = "1E8449"
_FG_RED   = "C0392B"
_FG_BLUE  = "1A5276"
_BRD      = "BFC9CA"


def _bd():
    s = Side(style="thin", color=_BRD)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)
def _fnt(sz=10, bold=False, color=_FG_D):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def _align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v)

def _L(ws, row, col, text):
    """Метка (тёмный текст, лёгкий фон)."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = _fnt(10, False, _FG_G)
    c.fill = _fill(_BG_INPUT)
    c.border = _bd()
    c.alignment = _align("left")

def _I(ws, row, col, value, fmt=None):
    """Input — редактируемое значение (голубой фон)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(10, True, _FG_BLUE)
    c.fill = _fill(_BG_INPUT)
    c.border = _bd()
    c.alignment = _align("right")
    if fmt:
        c.number_format = fmt
    return c

def _C(ws, row, col, formula_or_value, fmt=None):
    """Calculated — формула или значение (белый фон)."""
    c = ws.cell(row=row, column=col, value=formula_or_value)
    c.font = _fnt(10, False, _FG_D)
    c.fill = _fill(_BG_CALC)
    c.border = _bd()
    c.alignment = _align("right")
    if fmt:
        c.number_format = fmt
    return c

def _O(ws, row, col, formula_or_value, bold=True):
    """Output — итоговое значение (розовый фон)."""
    c = ws.cell(row=row, column=col, value=formula_or_value)
    c.font = Font(name="Calibri", size=10, bold=bold, color=_FG_D)
    c.fill = _fill(_BG_OUT)
    c.border = _bd()
    c.alignment = _align("right")
    return c

def _H(ws, row, col, text, end_col=None, bg=_BG_SECT, sz=10):
    """Заголовок секции."""
    if end_col and end_col > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=sz, bold=True, color=_FG_W)
    c.fill = _fill(bg)
    c.border = _bd()
    c.alignment = _align("left")

def _note(ws, row, col, text):
    """Аннотация справа (серый курсив)."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=9, italic=True, color=_FG_G)
    c.alignment = _align("left")


# Маппинг col-индекс → Excel-буква для col E..J (9 = I, но E=5)
# E=5, F=6, G=7, H=8, I=9, J=10, K=11, L=12, M=13
_YEAR_COLS = [5, 6, 7, 8, 9, 10]   # E=base, F..J=proj (5 years)
_PROJ_COLS = [6, 7, 8, 9, 10]       # F..J (projection years 1-5)


def _col(c: int) -> str:
    """1-based column index → Excel letter."""
    return get_column_letter(c)


def _build_dcf_sheet(ws, rec: dict, fund: Optional[dict]):
    """Заполнить один лист DCF для тикера rec."""

    ticker = rec.get("ticker", "X")
    name   = rec.get("name") or ticker
    cur    = rec.get("currency") or "USD"

    # ── Ширины ───────────────────────────────────────────────────────────────
    col_widths = {
        "A": 26, "B": 12, "C": 12, "D": 14,
        "E": 13, "F": 13, "G": 13, "H": 13, "I": 13, "J": 13,
        "K": 2,  "L": 24, "M": 14,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22

    # ── Row 1: шапка ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    c = ws.cell(row=1, column=1,
                value=f"DCF Valuation — {name}  ({ticker})  [{cur}]")
    c.font = Font(name="Calibri", size=13, bold=True, color=_FG_W)
    c.fill = _fill(_BG_DARK)
    c.alignment = _align("left")

    # ── Номера годов в строке 2 (col E..J) ───────────────────────────────────
    # E=0 (base), F=1..J=5
    for i, col in enumerate(_YEAR_COLS):
        c2 = ws.cell(row=2, column=col, value=i)
        c2.font = Font(name="Calibri", size=11, bold=True, color=_FG_W)
        c2.fill = _fill(_BG_YEAR)
        c2.border = _bd()
        c2.alignment = _align("right")
    # L1 = 5 (terminal year label)
    c_term = ws.cell(row=2, column=12, value=5)
    c_term.font = Font(name="Calibri", size=11, bold=True, color=_FG_W)
    c_term.fill = _fill(_BG_YEAR)

    # ── Извлекаем базовые данные из фундаменталов ────────────────────────────
    # Берём самый свежий год
    inc_years = sorted((fund or {}).get("income", {}).keys(), reverse=True)
    bal_years = sorted((fund or {}).get("balance", {}).keys(), reverse=True)
    cf_years  = sorted((fund or {}).get("cashflow", {}).keys(), reverse=True)

    def _get(src_key: str, yf_key: str, default=None):
        years = sorted((fund or {}).get(src_key, {}).keys(), reverse=True)
        for yr in years[:1]:
            v = (fund or {}).get(src_key, {}).get(yr, {}).get(yf_key)
            if v is not None:
                try: return float(v)
                except: pass
        return default

    rev_base   = _get("income",   "Total Revenue",           rec.get("market_cap", 1000) * 0.15)
    ebitda_mg  = None
    if rev_base:
        ebitda_v = _get("income", "EBITDA")
        if ebitda_v and rev_base:
            ebitda_mg = round(ebitda_v / rev_base, 4)
    if ebitda_mg is None:
        ebitda_mg = 0.45   # default 45%

    dna_base   = _get("income",   "Reconciled Depreciation") or _get("cashflow", "Depreciation And Amortization")
    dna_pct    = round(dna_base / rev_base, 4) if (dna_base and rev_base) else 0.03

    capex_base = _get("cashflow", "Capital Expenditure")
    capex_pct  = round(abs(capex_base) / rev_base, 4) if (capex_base and rev_base) else 0.08

    cash_eq    = _get("balance", "Cash And Cash Equivalents", 0)
    total_debt = _get("balance", "Total Debt",  rec.get("total_debt") or 0)
    shares     = _get("balance", "Share Issued") or _get("balance", "Ordinary Shares Number")
    if not shares:
        mktcap = rec.get("market_cap")
        price  = rec.get("price")
        shares = round(mktcap / price, 2) if (mktcap and price and price > 0) else 10.0
    else:
        shares = round(shares / 1e9, 4)   # переводим в миллиарды если нужно

    # ── СЕКЦИЯ: Income Statement ──────────────────────────────────────────────
    _H(ws, 3, 1, "income statement", end_col=10)   # A3

    # Row 4: revenue
    _L(ws, 4, 1, "revenue")
    _I(ws, 4, 5, round(rev_base / 1e6, 2) if rev_base else 1000.0, "#,##0.00")  # E4 = base (млн)
    for c_idx in _PROJ_COLS:
        prev = _col(c_idx - 1)
        _C(ws, 4, c_idx, f"={prev}4*(1+{_col(c_idx)}5)", "#,##0.00")

    _note(ws, 4, 12, "average growth rate")
    _O(ws, 4, 13, f"=AVERAGE(F5:J5)", "0.0%")

    # Row 5: %growth
    _L(ws, 5, 1, "%growth")
    growth_default = 0.10
    for c_idx in _PROJ_COLS:
        _I(ws, 5, c_idx, growth_default, "0%")

    # Row 6: (blank separator)

    # Row 7: ebitda
    _L(ws, 7, 1, "ebitda")
    for c_idx in _YEAR_COLS:
        _C(ws, 7, c_idx, f"={_col(c_idx)}4*{_col(c_idx)}8", "#,##0.00")

    # Row 8: % margin
    _L(ws, 8, 1, "% margin")
    for c_idx in _YEAR_COLS:
        _I(ws, 8, c_idx, ebitda_mg, "0%")

    # Row 9: (blank separator)

    # Row 10: d&a
    _L(ws, 10, 1, "d&a")
    _I(ws, 10, 5, round(dna_base / 1e6, 2) if dna_base else round(rev_base / 1e6 * dna_pct, 2), "#,##0.00")
    for c_idx in _PROJ_COLS:
        _C(ws, 10, c_idx, f"={_col(c_idx)}4*$M$10", "#,##0.00")
    _note(ws, 10, 12, "d&a % of revenue")
    _I(ws, 10, 13, dna_pct, "0%")

    # Row 11: ebit
    _L(ws, 11, 1, "ebit")
    for c_idx in _YEAR_COLS:
        _C(ws, 11, c_idx, f"={_col(c_idx)}7-{_col(c_idx)}10", "#,##0.00")

    # ── СЕКЦИЯ: Free Cash Flow Statement ────────────────────────────────────
    _H(ws, 14, 1, "free cash flow statement", end_col=10)  # A14

    # Row 15: ebitda (copy)
    _L(ws, 15, 1, "ebitda")
    for c_idx in _YEAR_COLS:
        _C(ws, 15, c_idx, f"={_col(c_idx)}7", "#,##0.00")

    # Row 16: tax
    _L(ws, 16, 1, "tax")
    for c_idx in _YEAR_COLS:
        _C(ws, 16, c_idx, f"=-{_col(c_idx)}11*$M$16", "#,##0.00")
    _note(ws, 16, 12, "tax rate")
    _I(ws, 16, 13, 0.20, "0%")

    # Row 17: capex
    _L(ws, 17, 1, "capex")
    for c_idx in _YEAR_COLS:
        _C(ws, 17, c_idx, f"=({_col(c_idx)}4*$M$17)*-1", "#,##0.00")
    _note(ws, 17, 12, "capex % of revenue")
    _I(ws, 17, 13, capex_pct, "0%")

    # Row 18: change in net wc
    _L(ws, 18, 1, "change in net wc")
    for c_idx in _YEAR_COLS:
        _C(ws, 18, c_idx, f"={_col(c_idx)}4*(-$M$18)", "#,##0.00")
    _note(ws, 18, 12, "(-) change in net wc % of revenue")
    _I(ws, 18, 13, 0.055, "0%")

    # Row 19: unlevered cash flow
    _L(ws, 19, 1, "unlevered cash flow")
    for c_idx in _YEAR_COLS:
        _C(ws, 19, c_idx, f"=SUM({_col(c_idx)}15:{_col(c_idx)}18)", "#,##0.00")

    # ── СЕКЦИЯ: Discount Cash Flow ────────────────────────────────────────────
    _H(ws, 22, 1, "discount cash flow", end_col=10)  # A22

    # Row 23: annual unlevered free cash flow
    _L(ws, 23, 1, "annual unlevered free cash flow")
    for c_idx in _YEAR_COLS:
        _C(ws, 23, c_idx, f"={_col(c_idx)}19", "#,##0.00")

    _note(ws, 23, 12, "terminal multiple")
    _I(ws, 23, 13, 5.0)

    # Row 24: terminal value (only col J)
    _L(ws, 24, 1, "terminal value")
    _note(ws, 21, 12, "perpetual grow rate")
    _I(ws, 21, 13, 0.04, "0%")
    _C(ws, 24, 10, "=J23*(1+$M$21)/(M27-M21)", "#,##0.00")  # J24

    # Row 25: total unlevered free cash flow
    _L(ws, 25, 1, "total unlevered free cash flow")
    for c_idx in _YEAR_COLS:
        _C(ws, 25, c_idx, f"=SUM({_col(c_idx)}23:{_col(c_idx)}24)", "#,##0.00")

    # Row 26-27: CAPM / WACC
    _note(ws, 26, 12, "capm")
    wacc_default = rec.get("region") == "KZ" and 0.14 or 0.094
    capm_default = wacc_default + 0.08
    _I(ws, 26, 13, round(capm_default, 3), "0%")
    _note(ws, 27, 12, "wacc")
    _I(ws, 27, 13, round(wacc_default, 3), "0%")

    # Row 28: present value ufcf
    _L(ws, 28, 1, "present value ufcf")
    _O(ws, 28, 4, "=SUM(F28:J28)", "#,##0.00")  # D28 = total
    # E28 = base year (не дисконтируется)
    _C(ws, 28, 5, "=E25", "#,##0.00")
    for c_idx in _PROJ_COLS:
        _C(ws, 28, c_idx, f"={_col(c_idx)}25/(1+$M$27)^{_col(c_idx)}2", "#,##0.00")

    # Row 29: net pv
    _L(ws, 29, 1, "net pv (NPV formula)")
    _O(ws, 29, 4, "=NPV(M27,F25:J25)", "#,##0.00")

    # ── СЕКЦИЯ: Bridge to Equity ──────────────────────────────────────────────
    _H(ws, 33, 1, "equity bridge", end_col=3)

    _L(ws, 34, 1, "sum of fcff")
    _O(ws, 34, 2, "=SUM(F28:J28)", "#,##0.00")

    _L(ws, 35, 1, "cash / cash equivalents")
    _I(ws, 35, 2, round(cash_eq / 1e6, 2) if cash_eq else 0, "#,##0.00")

    _L(ws, 36, 1, "total debt")
    _I(ws, 36, 2, round(abs(total_debt) / 1e6, 2) if total_debt else 0, "#,##0.00")

    _L(ws, 37, 1, "shares outstanding (M)")
    shares_val = shares if shares > 1000 else shares  # already in millions or billions?
    # if shares > 1000, assume they are in units, convert to millions
    if shares_val > 5000:
        shares_val = round(shares_val / 1e6, 2)
    _I(ws, 37, 2, round(shares_val, 4))

    _L(ws, 38, 1, "equity value")
    _O(ws, 38, 2, "=B34+B35-B36", "#,##0.00")

    _L(ws, 39, 1, "dcf price per share")
    _O(ws, 39, 2, "=B38/B37", "#,##0.00")

    # Current price for comparison
    if rec.get("price"):
        _L(ws, 40, 1, "current market price")
        _I(ws, 40, 2, round(float(rec["price"]), 2), "#,##0.00")
        _L(ws, 41, 1, "upside / downside")
        _O(ws, 41, 2, "=(B39-B40)/B40", "0.0%")

    # ── Легенда ───────────────────────────────────────────────────────────────
    ws.cell(row=43, column=1, value="Синий = входные данные (редактируемые)  |  Белый = расчётные  |  Розовый = итоговые"
            ).font = Font(name="Calibri", size=8, italic=True, color=_FG_G)


def build_dcf(record: dict, fund_data: Optional[dict] = None) -> bytes:
    """Построить DCF Excel для одного тикера. Возвращает bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (record.get("ticker") or "DCF")[:31]
    _build_dcf_sheet(ws, record, fund_data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_dcf_multi(records: list, fund_map: Optional[dict] = None) -> bytes:
    """Построить DCF Excel с несколькими листами (один лист = один тикер)."""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for rec in records:
        ws = wb.create_sheet(title=(rec.get("ticker") or "X")[:31])
        fd = (fund_map or {}).get(rec.get("ticker", ""))
        _build_dcf_sheet(ws, rec, fd)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
